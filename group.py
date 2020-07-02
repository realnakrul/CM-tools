from sys import argv
import tools
from pprint import pprint
from openpyxl import Workbook
from openpyxl import load_workbook


def main():
    if tools.check_input(argv[1]):
        source_file = argv[1]
        destination_file = f'grp_{source_file}'
        print(f'Output file: {destination_file}')
        print('-'*80)
        # matrix = pandas.read_excel(source_file)
        # print(matrix)
        source_book = load_workbook(source_file)
        destination_book = Workbook(write_only=True)
        for sheet in tools.select_sheet(source_book):
            print('-'*80)
            print(sheet.title)
            # print(sheet.max_column)
            # print(sheet.max_row)
            raw_matrix_list = tools.read_sheet(sheet)
            clean_matrix_list = tools.clean_list(raw_matrix_list)
            if not tools.consistency_check(clean_matrix_list):
                devices = tools.get_unique_values(clean_matrix_list, [0, 2])
                racks = tools.get_unique_values(clean_matrix_list, [6, 9])
                # pprint(clean_matrix_list)
                # tools.split_interfaces(clean_matrix_list, [0, 1])
                # split_list = tools.split_interfaces([0, 1], clean_matrix_list)
                # clean_matrix_list = tools.populate_b(clean_matrix_list)
                clean_matrix_list = tools.engineer_format(clean_matrix_list)
                clean_matrix_list = tools.group_by_device(devices, clean_matrix_list)
                destination_book = tools.add_to_sheet(destination_book, sheet.title, clean_matrix_list)
            else:
                print('Processing impossible, skipping')
        print('-'*80)
        print(f'Saving result workbook as {destination_file}')
        destination_book.save(destination_file)


if __name__ == '__main__':
    main()
