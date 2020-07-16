from os import path, rename
from collections import namedtuple, Counter
from itertools import combinations
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill


Link = namedtuple('Link', 'a_name, a_interface, b_name, b_interface, a_sfp, a_patch, a_rack, \
                           b_sfp, b_patch, b_rack, comment')


HEADERS = {'summary': ['Source Rack', 'Destination Rack', 'Link type', 'Amount'],
           'connectivity': ['Device A name', 'Device A port', 'Device B name', 'Device B port',
                            'Device A SFP', 'Device A patch cord', 'Device A rack',
                            'Device B SFP', 'Device B patch cord', 'Device B rack', 'Comment']}


def check_src_file(file):
    """
    Checks if source file exists
    :param file: Source file
    :return: Boolean
    """
    print('-'*80)
    result = path.exists(file)
    print(f"Checking if input file \'{file}\' exist: {result}")
    return result


def check_dst_file(file):
    """
    Checks if destination file is not open and available for writing
    :param file: Destination file name
    :return:
    """
    if path.exists(file):
        while True:
            try:
                rename(file, 'tmpfile.xlsx')
                rename('tmpfile.xlsx', file)
                break
            except OSError:
                print('Output file is open. Close the file and hit ENTER, please')
                input()


def select_sheet(book):
    sheets = []
    print('Available sheets: ')
    for ind, sheet in enumerate(book.sheetnames, 1):
        print(f'{ind}. {sheet}')
    print(f'{len(book.sheetnames)+1}. ALL')
    correct = False
    while not correct:
        try:
            sheet_ind = int(input(f'Select sheet (1 - {len(book.sheetnames) + 1}): '))
            if sheet_ind == len(book.sheetnames)+1:
                sheets = book.worksheets
                correct = True
            elif sheet_ind <= len(book.sheetnames)+1:
                sheets.append(book.worksheets[sheet_ind - 1])
                correct = True
        except ValueError:
            print('Select sheet number')
        except IndexError:
            print(f'Select sheet number in range (1 - {len(book.sheetnames)+1})')
        except Exception as e:
            print(type(e), e)
    return sheets


def read_sheet(sheet):
    matrix_list = []
    for row in sheet.iter_rows():
        row_list = []
        for cell in row:
            row_list.append(cell.value)
        matrix_list.append(row_list)
    return matrix_list


def clean_list(raw_matrix):
    """
    Removes all matrix entries without B device name (Device A open interfaces)
    :param raw_matrix: Connectivity matrix with Device A open interfaces and header
    :return: Connectivity matrix without header and Device A open interfaces
    """
    clean_matrix = []
    correct_input = False
    print('Cleaning connectivity matrix')
    print('Header row MUST be removed for processing (will be added back later)')
    print('First row:')
    print('| ', end='')
    for cell in raw_matrix[0]:
        print(cell, end=' | ')
    print()
    while not correct_input:
        answer = input('Is it header (y/n): ').lower()
        if answer == 'y':
            correct_input = True
            header = True
        elif answer == 'n':
            correct_input = True
            header = False
        else:
            print('y or n')
    print(f'\tRemoving header = {header}')
    if header:
        _ = raw_matrix.pop(0)
    for line in raw_matrix:
        link = Link(*line)
        if not link.a_rack:
            line[6] = 'Unknown'
        if not link.b_rack:
            line[9] = 'Unknown'
        if link.b_name:
            clean_matrix.append(line)
    print(f'\t{len(raw_matrix)-len(clean_matrix)} open interfaces removed')
    return clean_matrix


def group_by_device(devices: list, matrix: list, add_name=False):
    """
    Groups matrix rows by device name
    :param devices: Unique list of devices in the matrix
    :param matrix: Connectivity matrix in ENGINEER format
    :param add_name: Controls if device name required as a caption before group
    :return: Connectivity matrix grouped by device
    """
    result = []
    print('Grouping by device')
    print(f'\tSet device name as a header = {add_name}')
    for device in devices:
        device_group = []
        if add_name:
            result.append([device])
        for line in matrix:
            if device == str(line[0]):
                device_group.append(line)
        device_group.sort(key=lambda x: x[1])
        if device_group:
            result += device_group
    result.sort()
    return result


def add_to_sheet(book, sheet_name, data: list, header_type: str):
    """
    Adds excel sheet to workbook and fills it with data
    :param header_type: 'connectivity' or 'summary'
    :param book: Excel workbook
    :param sheet_name: Excel sheet name
    :param data: List to write to excel sheet
    :return: Excel workbook with given sheet
    """
    print(f'Adding {header_type} header')
    header = HEADERS[header_type]
    data.insert(0, header)
    print(f"Adding excel sheet \'{sheet_name}\' to result workbook")
    book.create_sheet(sheet_name)
    sheet = book[sheet_name]
    for line in data:
        sheet.append(line)
    return book


def get_unique_values(matrix: list, index_columns: list):
    """
    Gets unique values from a given matrix column
    :param matrix: Connectivity matrix
    :param index_columns: List of column indexes to get unique values
    :return: List of unique values
    """
    result = set()
    for i in index_columns:
        for line in matrix:
            result.add(str(line[i]).strip())
    return list(result)


def legacy_split_interfaces(matrix, device_columns: list):
    """
    Splits single list item with device and interface into two list items, assuming that interface is a substring after
    last space.
    :param device_columns: is a list of indexes - specifies which list items to split
    :param matrix: Connectivity matrix with device and interface as one cell
    """
    result = []
    for line in matrix:
        new_line = []
        for ind, cell in enumerate(line):
            if ind in device_columns:
                *device, interface = cell.split()
                new_line.extend([' '.join(device), interface])
            else:
                new_line.append(cell)
        result.append(new_line)
    return result


def legacy_populate_b(matrix):
    """
    Populates B device SFP, patch cord and rack from REVERSE record
    :param matrix: Clean connectivity matrix in FORWARD and REVERSE (Engineer) format without B SFP, Patch cord and rack
    :return: Connectivity matrix in FORWARD and REVERSE format with populated B SFP, Patch cord and rack
    """
    result = []
    for a_line in matrix:
        ab_line = []
        a_line_a_name = a_line[0]
        a_line_a_interface = a_line[1]
        a_line_b_name = a_line[2]
        a_line_b_interface = a_line[3]
        a_line_a_sfp = a_line[4]
        a_line_a_patch = a_line[5]
        a_line_a_rack = a_line[6]
        a_line_comment = a_line[7]
        for b_line in matrix:
            b_line_a_name = b_line[0]
            b_line_a_interface = b_line[1]
            b_line_a_sfp = b_line[4]
            b_line_a_patch = b_line[5]
            b_line_a_rack = b_line[6]
            if a_line_b_name == b_line_a_name and \
                    a_line_b_interface == b_line_a_interface:
                ab_line = [a_line_a_name, a_line_a_interface, a_line_b_name, a_line_b_interface,
                           a_line_a_sfp, a_line_a_patch, a_line_a_rack,
                           b_line_a_sfp, b_line_a_patch, b_line_a_rack, a_line_comment]
        if not ab_line:
            ab_line = [a_line_a_name, a_line_a_interface, a_line_b_name, a_line_b_interface,
                       a_line_a_sfp, a_line_a_patch, a_line_a_rack,
                       '', '', '', a_line_comment]
        result.append(ab_line)
    return result


def get_reverse(matrix: list, link_str: str):
    """
    Searches for REVERSE link for given link and matrix
    :param matrix: Clean connectivity matrix
    :param link_str: FORWARD link
    :return: REVERSE link
    """
    forward = Link(*link_str)
    result = ''
    for line in matrix:
        reverse = Link(*line)
        if forward.a_name == reverse.b_name and forward.b_name == reverse.a_name and \
                forward.a_interface == reverse.b_interface and forward.b_interface == reverse.a_interface:
            result = line
    return result


def engineer_format(matrix: list):
    """
    Creates ENGINEER format matrix from TECHNICIAN format
    :param: matrix: Clean connectivity matrix
    :return: Connectivity matrix in ENGINEER format
    """
    print('Enforcing ENGINEER format matrix')
    reverse_list = []
    result = []
    for line in matrix:
        forward = Link(*line)
        result.append(line)
        if not get_reverse(matrix, line):
            reverse_list.append([forward.b_name, forward.b_interface, forward.a_name, forward.a_interface,
                                 forward.b_sfp, forward.b_patch, forward.b_rack,
                                 forward.a_sfp, forward.a_patch, forward.a_rack, forward.comment])
            print('\tReverse fail:', line)
    print(f'\t{len(reverse_list)} reverse connections added')
    if reverse_list:
        result += reverse_list
    return result


def technician_format(matrix: list):
    """
    Creates TECHNICIAN format matrix from ENGINEER format
    :param: matrix: Clean connectivity matrix
    :return: Connectivity matrix in TECHNICIAN format
    """
    print('Enforcing TECHNICIAN format matrix')
    result = []
    for line in matrix:
        if not get_reverse(result, line):
            result.append(line)
    print(f'\t{len(matrix)-len(result)} reverse connections removed')
    return result


def consistency_check(matrix: list, devices: list):
    """
    Checks if interfaces are not duplicated or linked to itself
    :param devices: List of available devices
    :param matrix: Clean connectivity matrix
    :return: List of warning if detected
    """
    result = []
    for first_ind, first_line in enumerate(matrix):
        first_link = Link(*first_line)
        if first_link.a_name == first_link.b_name and first_link.a_interface == first_link.b_interface:
            warn = f'Warning: {first_link.a_name} {first_link.a_interface} links to itself.'
            print(f'\t{warn}')
            result.append(warn)
        if not first_link.a_patch and not first_link.b_patch:
            warn = f'Warning: {first_link.a_name} {first_link.a_interface} <-> {first_link.b_name} ' \
                   f'{first_link.b_interface} At least one patch cord should be filled.'
            print(f'\t{warn}')
            result.append(warn)
        if not first_link.a_interface or not first_link.b_interface:
            warn = f'Warning: {first_link.a_name} {first_link.a_interface} <-> {first_link.b_name} ' \
                   f'{first_link.b_interface} Interface can not be empty.'
            print(f'\t{warn}')
            result.append(warn)
        for second_ind, second_line in enumerate(matrix):
            second_link = Link(*second_line)
            if first_ind != second_ind and first_link.a_name == second_link.a_name and \
                    first_link.a_interface == second_link.a_interface:
                warn = f'Warning: {first_link.a_name} {first_link.a_interface} duplicated. -> {first_link.b_name} ' \
                       f'{first_link.b_interface} and {second_link.b_name} {second_link.b_interface}'
                print(f'\t{warn}')
                result.append(warn)
            elif first_ind != second_ind and first_link.b_name == second_link.b_name and \
                    first_link.b_interface == second_link.b_interface:
                warn = f'Warning: {first_link.b_name} {first_link.b_interface} duplicated. -> {first_link.a_name} ' \
                       f'{first_link.a_interface} and {second_link.a_name} {second_link.a_interface}'
                print(f'\t{warn}')
                result.append(warn)
    for device in devices:
        location = get_rack_by_device(matrix, device)
        if len(location) > 1:
            warn = f'Warning: {device} located in more then one rack: {", ".join(location)}'
            print(f'\t{warn}')
            result.append(warn)
    return result


def get_rack_by_device(matrix: list, device: str):
    """
    Gets device location racks
    :param device:
    :param matrix: Clean connectivity matrix
    :return: List of device location racks
    """
    result = set()
    for line in matrix:
        link = Link(*line)
        if link.a_name == device:
            result.add(link.a_rack)
        elif link.b_name == device:
            result.add(link.b_rack)
    return list(result)


def rack_to_rack_summary(racks: list, matrix: list):
    """
    Creates all possible combinations of existing racks and counts rack to rack connections
    :param racks: All racks list
    :param matrix: Connectivity matrix in ENGINEERING format
    :return: Summary list of rack to rack connections
    """
    print('Getting summary of rack to rack connections')
    result = []
    for racks_combination in list(combinations(racks, 2)):
        rack_to_rack_links = []
        for line in matrix:
            link = Link(*line)
            if racks_combination[0] == link.a_rack and racks_combination[1] == link.b_rack:
                rack_to_rack_links.append(line)
        if rack_to_rack_links:
            for type_sum in links_by_type_summary(rack_to_rack_links):
                result.append([racks_combination[0], racks_combination[1], type_sum[0], type_sum[1]])
    return result


def links_by_type_summary(matrix: list):
    """
    Creates links by type summary
    :param matrix: Connectivity matrix between two specific racks
    :return: List of connection types with connections amount
    """
    link_types = []
    for line in matrix:
        link = Link(*line)
        if link.a_patch:
            link_types.append(link.a_patch)
        elif link.b_patch:
            link_types.append(link.b_patch)
    result = [[key, value] for key, value in Counter(link_types).items()]
    return result


def add_filters(file_name):
    """
    Adds Excel filters to all sheets of workbook
    :param file_name: Excel workbook file name
    :return:
    """
    filter_book = load_workbook(file_name)
    for sheet_name in filter_book.sheetnames:
        sheet = filter_book[sheet_name]
        for col in range(1, sheet.max_column + 1):
            max_len = 0
            sheet[get_column_letter(col)+'1'].fill = PatternFill(start_color='FF036c9b', end_color='FF036c9b',
                                                                 fill_type='solid')
            sheet[get_column_letter(col) + '1'].font = Font(color='FFFFFFFF')
            for row in range(1, sheet.max_row + 1):
                cell = sheet[get_column_letter(col)+str(row)]
                if cell.value:
                    if len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
            sheet.column_dimensions[get_column_letter(col)].width = max_len + 3
        sheet.auto_filter.ref = sheet.dimensions
        sheet.freeze_panes = 'A2'
    filter_book.save(file_name)
