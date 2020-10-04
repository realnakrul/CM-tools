from sys import argv
import tools
from openpyxl import Workbook
from openpyxl import load_workbook


def main():
    if tools.check_src_file(argv[1]):
        src_file = argv[1]
        if len(argv) == 3:
            dst_file = argv[2]
        else:
            dst_file = tools.sel_dst_file(argv[1])
        print(f'Output file: {dst_file}')
        tools.check_dst_file(dst_file)
        print('-'*80)
        src_book = load_workbook(src_file)
        dst_book = Workbook(write_only=True)
        sheets = tools.select_sheet(src_book)
        print('-' * 80)
        for sheet in sheets:
            print('-'*80)
            print(f'Current worksheet: {sheet.title}')
            raw_matrix_list = tools.read_sheet(sheet)
            clean_matrix_list = tools.clean_list(raw_matrix_list)
            devices = tools.get_unique_values(clean_matrix_list, [0, 2])
            if not tools.consistency_check(clean_matrix_list, devices):
                racks = tools.get_unique_values(clean_matrix_list, [6, 9])
                clean_matrix_list = tools.group_by_device(devices, clean_matrix_list)
                engineer_matrix_list = tools.engineer_format(clean_matrix_list)
                technician_matrix_list = tools.technician_format(clean_matrix_list)
                summary_list = tools.rack_to_rack_summary(racks, engineer_matrix_list)
                dst_book = tools.add_to_sheet(dst_book, sheet.title + ' ENG', engineer_matrix_list, 'connectivity')
                dst_book = tools.add_to_sheet(dst_book, sheet.title + ' TEC', technician_matrix_list, 'connectivity')
                dst_book = tools.add_to_sheet(dst_book, sheet.title + ' SUM', summary_list, 'summary')
            else:
                print('Processing impossible, skipping')
        print('-'*80)
        if dst_book.sheetnames:
            print(f'Saving result workbook as {dst_file}')
            dst_book.save(dst_file)
            tools.add_filters(dst_file)


if __name__ == '__main__':
    main()
